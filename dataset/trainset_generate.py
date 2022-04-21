import os
import numpy as np
import multiprocessing
from openpyxl import load_workbook
from yuv_loader import load_yuv_seq


def yuv_l2_100_frames_extract():
    # excel load
    wb = load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
    ws = wb.active

    for seq_num in range(3, 200):
        # print(seq_num)
        seq_name = str(ws.cell(seq_num,1).value)
        seq_width = int(ws.cell(seq_num,3).value)
        seq_height = int(ws.cell(seq_num,4).value)
        # print(seq_name, seq_width, seq_height)
        os.system('ffmpeg -s {}x{} -i {} -vf scale={}:{}:sws_flags=lanczos {}'.format(seq_width, seq_height, '/workspace/shared/YUV-HIF/yuv/'+seq_name+'.yuv', seq_width//2, seq_height//2, '/workspace/shared/YUV-HIF/yuv_l2_lanczos/'+seq_name+'.yuv'))
        seq_y, seq_u, seq_v = load_yuv_seq(seq_path='/workspace/shared/YUV-HIF/yuv_l2_lanczos/'+seq_name+'.yuv', h=int(seq_height//2), w=int(seq_width//2), tot_frm=100, bit=8)
        if not os.path.exists('/workspace/shared/YUV-HIF/l2_lanczos_100_np_y/'+seq_name):
            os.makedirs('/workspace/shared/YUV-HIF/l2_lanczos_100_np_y/'+seq_name)
        for num in range(100):
            frame_y = seq_y[num, :, :]
            np.save('/workspace/shared/YUV-HIF/l2_lanczos_100_np_y/'+seq_name+'/'+str(num+1)+'_y',frame_y)


def yuv_100_frames_extract():
    # excel load
    data = xlrd.open_workbook('/data/YUV-HIF/HIF-Database.xlsx')
    data.sheet_names()
    # print('sheets: ', str(data.sheet_names()))
    # tabel = data.sheet_by_name('sheet1')
    table = data.sheet_by_index(0)
    # print("总行数：", str(table.nrows))
    # print("总列数：", str(table.ncols))

    for seq_num in range(2, 199):
        print(seq_num)
        seq_name = str(table.cell(seq_num,0).value)
        seq_width = int(table.cell(seq_num,2).value)
        seq_height = int(table.cell(seq_num,3).value)
        print(seq_name)
        seq_y, seq_u, seq_v = load_yuv_seq(seq_path='/data/YUV-HIF/yuv/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
        if not os.path.exists('/data/YUV-HIF/org_100_np_y/'+seq_name):
            os.makedirs('/data/YUV-HIF/org_100_np_y/'+seq_name)
        for num in range(100):
            frame_y = seq_y[num, :, :]
            np.save('/data/YUV-HIF/org_100_np_y/'+seq_name+'/'+str(num+1)+'_y',frame_y)


def vvc_ra_l2_multiprocess(item):
    qp = 27
    wb = load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
    ws = wb.active

    seq_name = str(ws.cell(item,1).value)
    seq_width = int(ws.cell(item,3).value)//2
    seq_height = int(ws.cell(item,4).value)//2

    if os.path.exists('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name):
        return

    if not os.path.exists('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name)
    os.system('/workspace/lm/CRS/codes/dataset/VVCSoftware_VTM-VTM-10.0/bin/EncoderAppStatic -i {} -wdt {} -hgt {} -fr 25 -c /workspace/lm/CRS/codes/dataset/VVCSoftware_VTM-VTM-10.0/cfg/encoder_randomaccess_vtm.cfg --ConformanceWindowMode=1 --OutputBitDepth=8 -f 100 -q {} -o {}'.format('/workspace/shared/YUV-HIF/yuv_l2_lanczos/'+seq_name+'.yuv', seq_width, seq_height, qp, '/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name+'/'+seq_name+'.yuv'))
    
    if not os.path.exists('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
    for num in range(100):
        np.save('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def vvc_ra_org_multiprocess(item):
    qp = 47
    data = xlrd.open_workbook('/data/YUV-HIF/HIF-Database.xlsx')
    data.sheet_names()
    table = data.sheet_by_index(0)

    seq_name = str(table.cell(item,0).value)
    seq_width = int(table.cell(item,2).value)
    seq_height = int(table.cell(item,3).value)

    if os.path.exists('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv/'+seq_name):
        return

    if not os.path.exists('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv/'+seq_name):
        os.makedirs('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv/'+seq_name)
    os.system('/workspace/VVCSoftware_VTM-VTM-10.0/bin/EncoderAppStatic -i {} -wdt {} -hgt {} -fr 25 -c /workspace/VVCSoftware_VTM-VTM-10.0/cfg/encoder_randomaccess_vtm.cfg --OutputBitDepth=8 -f 100 -q {} -o {}'.format('/data/YUV-HIF/yuv/'+seq_name+'.yuv', seq_width, seq_height, qp, '/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv'))
    
    if not os.path.exists('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_y/'+seq_name):
        os.makedirs('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
    for num in range(100):
        np.save('/data/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def vvc_ldp_l2_multiprocess(item):
    qp = 42
    wb = load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
    ws = wb.active

    seq_name = str(ws.cell(item,1).value)
    seq_width = int(ws.cell(item,3).value)//2
    seq_height = int(ws.cell(item,4).value)//2

    if os.path.exists('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name):
        return

    if not os.path.exists('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name)
    os.system('/workspace/lm/CRS/codes/dataset/VVCSoftware_VTM-VTM-10.0/bin/EncoderAppStatic -i {} -wdt {} -hgt {} -fr 25 -c /workspace/lm/CRS/codes/dataset/VVCSoftware_VTM-VTM-10.0/cfg/encoder_lowdelay_P_vtm.cfg --ConformanceWindowMode=1 --OutputBitDepth=8 -f 100 -q {} -o {}'.format('/workspace/shared/YUV-HIF/yuv_l2_lanczos/'+seq_name+'.yuv', seq_width, seq_height, qp, '/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name+'/'+seq_name+'.yuv'))
    
    if not os.path.exists('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv_l2_lanczos/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
    for num in range(100):
        np.save('/workspace/shared/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_l2_lanczos_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def vvc_ldp_org_multiprocess(item):
    qp = 27
    data = xlrd.open_workbook('/data/YUV-HIF/HIF-Database.xlsx')
    data.sheet_names()
    table = data.sheet_by_index(0)

    seq_name = str(table.cell(item,0).value)
    seq_width = int(table.cell(item,2).value)
    seq_height = int(table.cell(item,3).value)

    if os.path.exists('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv/'+seq_name):
        return

    if not os.path.exists('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv/'+seq_name):
        os.makedirs('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv/'+seq_name)
    os.system('/workspace/VVCSoftware_VTM-VTM-10.0/bin/EncoderAppStatic -i {} -wdt {} -hgt {} -fr 24 -c /workspace/VVCSoftware_VTM-VTM-10.0/cfg/train_ldp.cfg --OutputBitDepth=8 -f 100 -q {} -o {}'.format('/data/YUV-HIF/yuv/'+seq_name+'.yuv', seq_width, seq_height, qp, '/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv'))
    
    if not os.path.exists('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_y/'+seq_name):
        os.makedirs('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
    for num in range(100):
        np.save('/data/YUV-HIF/vvc/LDP/qp'+str(qp)+'/np_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def av1_ldp_l2_multiprocess(item):
    qp = 32
    wb = load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
    ws = wb.active

    seq_name = str(ws.cell(item,1).value)
    seq_width = int(ws.cell(item,3).value)//2
    seq_height = int(ws.cell(item,4).value)//2

    if os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name):
        return

    if not os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name)
    os.system('/workspace/lm/CRS/codes/dataset/aom_build/aomenc --width={} --height={} --fps=30/1 --codec=av1 --test-decode=0 --cpu-used=0 --profile=0 --drop-frame=0 --static-thresh=0 --sharpness=0 --frame-parallel=0 --tile-columns=0 --kf-max-dist=100 --kf-min-dist=0 --end-usage=q --psnr -v --limit=100 --enable-cdef=1 --enable-restoration=1 --cq-level={} -o {} {}'.format(seq_width, seq_height, qp, '/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name+'/'+seq_name+'.ivf', '/workspace/shared/YUV-HIF/yuv_l2/'+seq_name+'.yuv'))
    os.system('/workspace/lm/CRS/codes/dataset/aom_build/aomdec -o {} {}'.format('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name+'/'+seq_name+'.yuv', '/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name+'/'+seq_name+'.ivf'))

    if not os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_l2_y/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_l2_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv_l2/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=100, bit=8)
    for num in range(100):
        np.save('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_l2_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def av1_ldp_org_multiprocess(item):
    qp = 42
    wb = load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
    ws = wb.active

    seq_name = str(ws.cell(item,1).value)
    seq_width = int(ws.cell(item,3).value)
    seq_height = int(ws.cell(item,4).value)

    if os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name):
        return

    if not os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name)
    os.system('/workspace/lm/CRS/codes/dataset/aom_build/aomenc --width={} --height={} --fps=30/1 --codec=av1 --test-decode=0 --cpu-used=0 --profile=0 --drop-frame=0 --static-thresh=0 --sharpness=0 --frame-parallel=0 --tile-columns=0 --kf-max-dist=10 --kf-min-dist=0 --end-usage=q --psnr -v --limit=50 --enable-cdef=1 --enable-restoration=1 --cq-level={} -o {} {}'.format(seq_width, seq_height, qp, '/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.ivf', '/workspace/shared/YUV-HIF/yuv/'+seq_name+'.yuv'))
    os.system('/workspace/lm/CRS/codes/dataset/aom_build/aomdec -o {} {}'.format('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv', '/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.ivf'))

    if not os.path.exists('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_y/'+seq_name):
        os.makedirs('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_y/'+seq_name)
    frame_y, frame_u, frame_v = load_yuv_seq(seq_path='/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/yuv/'+seq_name+'/'+seq_name+'.yuv', h=seq_height, w=seq_width, tot_frm=50, bit=8)
    for num in range(50):
        np.save('/workspace/shared/YUV-HIF/av1/LDP/qp'+str(qp)+'/np_y/'+seq_name+'/'+str(num+1)+'_y',frame_y[num, :, :])


def data_generate(opt):
    if opt == 'vvc_ra_l2_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(16)
        p.map(vvc_ra_l2_multiprocess, items)
        p.close()
        p.join()

    if opt == 'vvc_ldp_l2_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(16)
        p.map(vvc_ldp_l2_multiprocess, items)
        p.close()
        p.join()

    if opt == 'vvc_ldp_org_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(16)
        p.map(vvc_ldp_org_multiprocess, items)
        p.close()
        p.join()

    if opt == 'vvc_ld_l2_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(16)
        p.map(vvc_ld_l2_multiprocess, items)
        p.close()
        p.join()

    if opt == 'vvc_ld_org_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(16)
        p.map(vvc_ld_org_multiprocess, items)
        p.close()
        p.join()

    if opt == 'av1_ldp_l2_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(4)
        p.map(av1_ldp_l2_multiprocess, items)
        p.close()
        p.join()

    if opt == 'av1_ldp_org_multiprocess':
        items = []
        for i in range(3,200):
            items.append(i)
        p = multiprocessing.Pool(4)
        p.map(av1_ldp_org_multiprocess, items)
        p.close()
        p.join()

    
def file_examine():
    for qp in [27, 32, 37, 42]:
        print(qp, len(os.listdir('/workspace/shared/YUV-HIF/vvc/RA/qp'+str(qp)+'/np_l2_lanczos_y')))


if __name__ == '__main__':
    # yuv_l2_100_frames_extract()
    # yuv_100_frames_extract()
    data_generate('av1_ldp_l2_multiprocess')
    # data_generate('vvc_ldp_l2_multiprocess')
    # file_examine()