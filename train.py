import os
import random
import numpy as np
import openpyxl
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import Dataset
from torch.utils.data import DataLoader
from modules import crs

os.environ['CUDA_VISIBLE_DEVICES'] = '1'

num_epoch = 100

class ImagesFromFolder(Dataset):
    def __init__(self):
        self.images_list = []
        self.num = 0
        self.height = 128
        self.width = 128
        tmf_img_path = '/workspace/shared/YUV-HIF/vvc/RA/qp32/np_l2_lanczos_y/'
        stf_img_path = '/workspace/shared/YUV-HIF/vvc/RA/qp37/np_y/'
        gt_img_path = '/workspace/shared/YUV-HIF/org_100_np_y/'

        # tmf_img_path = '/workspace/shared/YUV-HIF/vvc/LDP/qp27/np_l2_lanczos_y/'
        # stf_img_path = '/workspace/shared/YUV-HIF/vvc/LDP/qp32/np_y/'
        # gt_img_path = '/workspace/shared/YUV-HIF/org_100_np_y/'

        wb = openpyxl.load_workbook('/workspace/shared/YUV-HIF/HIF-Database.xlsx')
        ws = wb.active

        for seq_num in range(3, 200):
            seq_name = str(ws.cell(seq_num,1).value)
            seq_width = int(ws.cell(seq_num,3).value)
            seq_height = int(ws.cell(seq_num,4).value)

            if seq_width < 2*self.width or seq_height < 2*self.height:
                continue

            # RA
            for num in range(3,96):
                if num % 32 == 1:
                    continue
                else:
                    in_y = tmf_img_path + seq_name + '/' + str(num) + '_y.npy'
                    ref1_y = tmf_img_path + seq_name + '/' + str(num-1) + '_y.npy'
                    ref2_y = tmf_img_path + seq_name + '/' + str(num+1) + '_y.npy'
                    intraf_y = stf_img_path + seq_name + '/' + str((num//32)*32+1) + '_y.npy'
                    intrab_y = stf_img_path + seq_name + '/' + str((num//32+1)*32+1) + '_y.npy'
                    # intra_y = gt_img_path + seq + '/1_y.npy'
                    gt_y = gt_img_path + seq_name + '/' + str(num) + '_y.npy'
                    self.images_list += [[in_y, ref1_y, ref2_y, intraf_y, intrab_y, gt_y]]
                    self.num = self.num + 1

            # # LDP
            # for num in range(2,100):
            #     # if num % 8 == 1:
            #     #     continue
            #     # else:
            #     in_y = tmf_img_path + seq_name + '/' + str(num) + '_y.npy'
            #     ref1_y = tmf_img_path + seq_name + '/' + str(num-1) + '_y.npy'
            #     ref2_y = tmf_img_path + seq_name + '/' + str(num+1) + '_y.npy'
            #     intraf_y = stf_img_path + seq_name + '/1_y.npy'
            #     # intrab_y = stf_img_path + seq_name + '/' + str((num//8)*8+2) + '_y.npy'
            #     # intra_y = gt_img_path + seq + '/1_y.npy'
            #     gt_y = gt_img_path + seq_name + '/' + str(num) + '_y.npy'
            #     self.images_list += [[in_y, ref1_y, ref2_y, intraf_y, gt_y]]
            #     self.num = self.num + 1
				
        print('train data:{}', self.num)

    def __getitem__(self, index):
        in_y = np.load(self.images_list[index][0]).astype(np.float32)/255.
        H, W = in_y.shape
        ref1_y = np.load(self.images_list[index][1]).astype(np.float32)/255.
        ref2_y = np.load(self.images_list[index][2]).astype(np.float32)/255.
        intraf_y = np.load(self.images_list[index][3]).astype(np.float32)/255.
        intrab_y = np.load(self.images_list[index][4]).astype(np.float32)/255.
        gt_y = np.load(self.images_list[index][5]).astype(np.float32)/255.

        tran_h = np.random.randint(0, H-self.height)
        tran_w = np.random.randint(0, W-self.width)

        k = np.random.randint(0, 4)
        if (np.random.randint(0, 2) == 1):
            in_y = torch.from_numpy(np.fliplr(np.rot90(in_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).copy()).unsqueeze(0)
            ref1_y = torch.from_numpy(np.fliplr(np.rot90(ref1_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).copy()).unsqueeze(0)
            ref2_y = torch.from_numpy(np.fliplr(np.rot90(ref2_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).copy()).unsqueeze(0)
            intraf_y = torch.from_numpy(np.fliplr(np.rot90(intraf_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width].copy(), k).copy()).copy()).unsqueeze(0)
            intrab_y = torch.from_numpy(np.fliplr(np.rot90(intrab_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width].copy(), k).copy()).copy()).unsqueeze(0)
            gt_y = torch.from_numpy(np.fliplr(np.rot90(gt_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width].copy(), k).copy()).copy()).unsqueeze(0)
        else:
            in_y = torch.from_numpy(np.rot90(in_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).unsqueeze(0)
            ref1_y = torch.from_numpy(np.rot90(ref1_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).unsqueeze(0)
            ref2_y = torch.from_numpy(np.rot90(ref2_y[tran_h:tran_h+self.height, tran_w:tran_w+self.width].copy(), k).copy()).unsqueeze(0)
            intraf_y = torch.from_numpy(np.rot90(intraf_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width], k).copy()).unsqueeze(0)
            intrab_y = torch.from_numpy(np.rot90(intrab_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width], k).copy()).unsqueeze(0)
            gt_y = torch.from_numpy(np.rot90(gt_y[2*tran_h:2*tran_h+2*self.height, 2*tran_w:2*tran_w+2*self.width], k).copy()).unsqueeze(0)

        sample = {'in_y': in_y, 'ref1_y': ref1_y, 'ref2_y': ref2_y, 'intraf_y': intraf_y, 'intrab_y': intrab_y, 'gt_y': gt_y}
        return sample

    def __len__(self):
        return self.num

def train():
    print('===> Loading data')
    train_data = ImagesFromFolder()
    train_loader = DataLoader(train_data, batch_size=4, shuffle=True)
    print('===> Building model')
    model = crs.CRS(num_in_ch=1, 
                    num_out_ch=1, 
                    num_feat=64, 
                    num_frame=3, 
                    num_extract_block=8, 
                    num_deformable_group=8, 
                    num_reconstruct_block=8, 
                    num_fusion_block=8, 
                    center_frame_idx=1)

    criterion = nn.L1Loss()

    print('===> Loading model')
    model.load_state_dict(torch.load('/workspace/lm/CRS/checkpoints/RA/qp32/epoch_22_loss_0.0139258.pkl'))
    # model.load_state_dict(torch.load('/checkpoints/ssr/hm/qp27/epoch_46_loss_0.0120599.pkl'), strict=False)

    print('===> Setting GPU')
    if torch.cuda.is_available():
        print('===>GPU used')
        model = model.cuda()
        # model = nn.DataParallel(model, device_ids=[0,1])

    print('===> Setting optimizer')
    opt = torch.optim.Adam(model.parameters(), lr=0.0001)
    
    print('===> Training')
    model.train()
    for epoch in range(23,num_epoch):
        sum_loss = 0.0
        for iteration, sample in enumerate(train_loader):
            if torch.cuda.is_available():
                in_y = sample['in_y'].cuda()
                ref1_y = sample['ref1_y'].cuda()
                ref2_y = sample['ref2_y'].cuda()
                intraf_y = sample['intraf_y'].cuda()
                intrab_y = sample['intrab_y'].cuda()
                gt_y = sample['gt_y'].cuda()
                frames_y = torch.stack([ref1_y, in_y, ref2_y], dim=1)

            opt.zero_grad()

            output_y = model(frames_y, intraf_y, intrab_y, train=True)

            rec_loss = criterion(output_y, gt_y)

            rec_loss.backward()
            opt.step()
            sum_loss += rec_loss.item()

            if iteration % 10 == 0:
                print("===> Epoch[{}]({}/{}): rec_loss:{:.7f}".format(epoch, iteration, len(train_loader), rec_loss.item()))
                # torch.save(model.state_dict(), '/checkpoints/ttvsr2/hm/qp37/{:.7f}.pkl'.format(rec_loss.item()))
            # if (iteration != 0) and (iteration % 500 == 0):
                # torch.save(model.state_dict(), '/checkpoints/ttvsr2/hm/qp37/{:.7f}.pkl'.format(rec_loss.item()))
        
        # save to checkpoint
        checkpoint_path = ('/workspace/lm/CRS/checkpoints/RA/qp32/epoch_{}_loss_{:.7f}.pkl'.format(epoch, sum_loss/len(train_loader)))
        torch.save(model.state_dict(), checkpoint_path)
        # torch.save(model.module.state_dict(), checkpoint_path)
        print('===> Model saved to {}'.format(checkpoint_path))

if __name__ == '__main__':
	train()

